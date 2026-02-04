import streamlit as st
import google.generativeai as genai
import json
import io
import time
import gc
import httpx
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Sayfa Ayarları
st.set_page_config(
    page_title="İSG Risk Değerlendirme Asistanı",
    page_icon="🛡️",
    layout="wide"
)

# === CUSTOM CSS - GÖRSEL ŞÖLEN ===
st.markdown("""
<style>
    /* Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700;800&display=swap');
    
    /* Global Styles */
    * {
        font-family: 'Inter', sans-serif;
    }
    
    /* Main Background with Gradient */
    .stApp {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        background-attachment: fixed;
    }
    
    /* Content Container with Glassmorphism */
    .main .block-container {
        background: rgba(255, 255, 255, 0.95);
        backdrop-filter: blur(10px);
        border-radius: 20px;
        padding: 3rem 2rem;
        box-shadow: 0 8px 32px 0 rgba(31, 38, 135, 0.37);
        border: 1px solid rgba(255, 255, 255, 0.18);
        animation: fadeIn 0.6s ease-in;
    }
    
    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(20px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    /* Headers */
    h1 {
        color: #1a202c;
        font-weight: 800;
        font-size: 3rem !important;
        margin-bottom: 0.5rem;
        animation: slideInDown 0.6s ease-out;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.1);
    }
    
    @keyframes slideInDown {
        from { opacity: 0; transform: translateY(-30px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    h2, h3 {
        color: #1a202c;
        font-weight: 700;
    }
    
    /* Form Containers */
    .stForm {
        background: linear-gradient(135deg, rgba(255,255,255,0.9) 0%, rgba(255,255,255,0.7) 100%);
        border-radius: 15px;
        padding: 2rem;
        box-shadow: 0 4px 20px rgba(0,0,0,0.1);
        border: 2px solid rgba(102, 126, 234, 0.2);
        transition: all 0.3s ease;
    }
    
    .stForm:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 30px rgba(102, 126, 234, 0.3);
    }
    
    /* Inputs */
    .stTextInput input, .stTextArea textarea {
        border-radius: 10px;
        border: 2px solid #e2e8f0;
        padding: 0.75rem;
        font-size: 1rem;
        transition: all 0.3s ease;
    }
    
    .stTextInput input:focus, .stTextArea textarea:focus {
        border-color: #667eea;
        box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
    }
    
    /* Buttons */
    .stButton button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        font-size: 1.1rem;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
        cursor: pointer;
    }
    
    .stButton button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6);
    }
    
    .stButton button:active {
        transform: translateY(0);
    }
    
    .stDownloadButton button {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(245, 87, 108, 0.4);
    }
    
    .stDownloadButton button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(245, 87, 108, 0.6);
    }
    
    /* Progress Bar */
    .stProgress > div > div {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 50%, #f093fb 100%);
        border-radius: 10px;
        height: 12px;
    }
    
    /* Slider */
    .stSlider > div > div > div {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
    
    /* Success/Error Messages */
    .stSuccess {
        background: linear-gradient(135deg, #84fab0 0%, #8fd3f4 100%);
        border-radius: 10px;
        padding: 1rem;
        border-left: 4px solid #84fab0;
    }
    
    .stError {
        background: linear-gradient(135deg, #fa709a 0%, #fee140 100%);
        border-radius: 10px;
        padding: 1rem;
        border-left: 4px solid #fa709a;
    }
    
    /* Expander */
    .streamlit-expanderHeader {
        background: linear-gradient(135deg, rgba(102, 126, 234, 0.1) 0%, rgba(118, 75, 162, 0.1) 100%);
        border-radius: 10px;
        font-weight: 600;
    }
    
    /* Footer Enhancement */
    .custom-footer {
        text-align: center;
        margin-top: 3rem;
        padding: 2rem;
    }
    
    .custom-footer h3 {
        color: #667eea;
        font-weight: 700;
        font-size: 1.3rem;
        margin-bottom: 0.5rem;
    }
    
    .custom-footer p {
        color: #4a5568;
        font-size: 0.95rem;
    }
    
    /* Hide Streamlit Branding */
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# === YARDIMCI FONKSİYONLAR ===

# Fine Kinney Risk Seviyeleri
RISK_LEVELS = {
    "tolerans_gosterilemez": {"min": 400, "max": float('inf'), "color": "FFFF0000", "label": "Tolerans Gösterilemez Risk"},
    "esasli": {"min": 200, "max": 400, "color": "FF808080", "label": "Esaslı Risk"},
    "onemli": {"min": 70, "max": 200, "color": "FF0070C0", "label": "Önemli Risk"},
    "olasi": {"min": 20, "max": 70, "color": "FFFFFF00", "label": "Olası Risk"},
    "onemsiz": {"min": 0, "max": 20, "color": "FF00B050", "label": "Önemsiz Risk"}
}

def get_risk_level(score):
    if score > 400: return RISK_LEVELS["tolerans_gosterilemez"]
    elif score > 200: return RISK_LEVELS["esasli"]
    elif score > 70: return RISK_LEVELS["onemli"]
    elif score > 20: return RISK_LEVELS["olasi"]
    else: return RISK_LEVELS["onemsiz"]

def create_excel(risk_data, workplace):
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Değerlendirme"
    
    headers = [
        "Sıra No", "Faaliyet Alanı", "Faaliyet Türü", 
        "Tehlike Tanımı", "Risk Tanımı (Olası Etki)",
        "O", "F", "Ş", "R", "Riskin Tanımı",
        "Planlanan Faaliyetler / DÖF", "Sorumlu", "Süre",
        "Sonraki O", "Sonraki F", "Sonraki Ş", "Sonraki R", "Sonraki Riskin Tanımı"
    ]
    ws.append(headers)
    
    # Stiller
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for item in risk_data:
        o = float(item.get('olasilik', 1))
        f = float(item.get('frekans', 1))
        s = float(item.get('siddet', 1))
        current_score = o * f * s
        current_level = get_risk_level(current_score)
        
        so = float(item.get('sonraki_olasilik', 0.2))
        sf = float(item.get('sonraki_frekans', 1))
        ss = float(item.get('sonraki_siddet', 1))
        next_score = so * sf * ss
        next_level = get_risk_level(next_score)
        
        # Önlemler listesini metne çevir
        onlemler = item.get('onlemler', '')
        if isinstance(onlemler, list):
            onlemler = '\n'.join([f"• {o}" for o in onlemler])
            
        row = [
            item.get('sira_no'), item.get('faaliyet_alani'), item.get('faaliyet_turu'),
            item.get('tehlike_tanimi'), item.get('risk_tanimi'),
            o, f, s, current_score, current_level["label"],
            onlemler, item.get('sorumlu'), item.get('sure'),
            so, sf, ss, next_score, next_level["label"]
        ]
        ws.append(row)
    
    # Hücre stilleri
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Mevcut Risk Rengi
        score_cell = ws.cell(row=row_idx, column=9)
        try:
            val = float(score_cell.value)
            level = get_risk_level(val)
            score_cell.fill = PatternFill(start_color=level["color"], end_color=level["color"], fill_type="solid")
            if level["color"] in ["FF0070C0", "FF808080", "FFFF0000"]:
                score_cell.font = Font(color="FFFFFF", bold=True)
        except: pass
        
        # Sonraki Risk Rengi
        next_score_cell = ws.cell(row=row_idx, column=17)
        try:
            val = float(next_score_cell.value)
            level = get_risk_level(val)
            next_score_cell.fill = PatternFill(start_color=level["color"], end_color=level["color"], fill_type="solid")
        except: pass

    # Sütun genişlikleri
    widths = {'A': 8, 'B': 18, 'C': 18, 'D': 35, 'E': 30, 'J': 22, 'K': 50, 'L': 25, 'R': 22}
    for col, width in widths.items(): ws.column_dimensions[col].width = width
    
    return wb

def fetch_risks_in_batches(api_key, model_name, workplace, total_items=50, batch_size=10, progress_bar=None, status_text=None):
    all_risks = []
    
    # Gemini Ayarlari
    genai.configure(api_key=api_key)
    # Seçilen modeli kullan
    model = genai.GenerativeModel(model_name)
    
    num_batches = (total_items + batch_size - 1) // batch_size
    
    for i in range(num_batches):
        start_idx = i * batch_size + 1
        current_batch_size = min(batch_size, total_items - len(all_risks))
        
        if status_text:
            status_text.text(f"⏳ RİSK ANALİZİ OLUŞTURULUYOR... (Paket {i+1}/{num_batches})")
        if progress_bar:
            progress_bar.progress((i) / num_batches)
            
        prompt = f"""
        Sen uzman bir İSG (İş Sağlığı ve Güvenliği) mühendisisin.
        Görev: '{workplace}' işyeri/sektörü için {current_batch_size} adet detaylı risk değerlendirmesi yap.
        ÖNEMLİ: Bu bir serinin parçasıdır. Risk numaraları {start_idx}'den başlayarak {start_idx + current_batch_size - 1}'e kadar gitmeli.

        Fine Kinney Metodu değerleri:
        - Olasılık (O): 0.2, 0.5, 1, 3, 6, 10
        - Frekans (F): 0.5, 1, 2, 3, 6, 10
        - Şiddet (Ş): 1, 3, 7, 15, 40, 100
        
        Çıktı formatı: Sadece saf JSON array döndür. Markdown bloğu kullanma.
        Her obje şu anahtarları içermeli:
        - sira_no (Integer: {start_idx} - {start_idx + current_batch_size - 1})
        - faaliyet_alani (Örn: Genel Yönetim, Üretim Alanı)
        - faaliyet_turu (Örn: Çalışma Ortamı, Makine Kullanımı)
        - tehlike_tanimi (Detaylı tehlike açıklaması)
        - risk_tanimi (Olası etki: yaralanma, ölüm, maddi hasar)
        - olasilik (Fine Kinney değeri)
        - frekans (Fine Kinney değeri)
        - siddet (Fine Kinney değeri)
        - onlemler (DÖF - Düzeltici/Önleyici Faaliyetler, maddeler halinde)
        - sorumlu (Örn: İşveren & İSG Uzmanı)
        - sure (Aksiyon süresi: "Hemen", "1 Ay" vb.)
        - sonraki_olasilik (DÖF sonrası)
        - sonraki_frekans (DÖF sonrası)
        - sonraki_siddet (DÖF sonrası)
        
        KRİTİK KURALLAR:
        1. DÖF sonrası Risk Skoru (O×F×Ş) KESİNLİKLE 70 veya altında olmalı.
        2. "{workplace}" sektörüne özel gerçekçi riskler üret.
        3. En az 1 tane yüksek (400+) risk olsun.
        """
        
        try:
            # count_tokens ile maliyet kontrolü yapılabilir ama şimdilik direkt generate_content
            response = model.generate_content(
                prompt,
                generation_config=genai.types.GenerationConfig(
                    temperature=0.7,
                    response_mime_type="application/json"
                )
            )
            
            content = response.text.strip()
            # Bazı durumlarda yine de md block gelebilir
            if content.startswith("```json"): content = content[7:]
            if content.startswith("```"): content = content[3:]
            if content.endswith("```"): content = content[:-3]
            
            batch_data = json.loads(content)
            if isinstance(batch_data, dict): batch_data = [batch_data]
            all_risks.extend(batch_data)
            
            # Bellek Temizliği gerekmez ama yine de
            del content, response
            gc.collect()
            
        except Exception as e:
            st.error(f"Paket {i+1} Hatası: {str(e)}")
            time.sleep(2) # Hata durumunda bekle
            continue

    if progress_bar: progress_bar.progress(1.0)
    return all_risks

# === ARAYÜZ ===
col1, col2 = st.columns([3, 1])
with col1:
    st.title("⚡ İş Güvenliği Risk Analizi")
    st.markdown("""
        <p style='font-size: 1.2rem; color: #4a5568; margin-top: -1rem;'>
            🚀 <b>Yapay zeka destekli</b> otomatik risk değerlendirmesi<br>
            ✨ Dakikalar içinde profesyonel Excel raporları oluşturun
        </p>
    """, unsafe_allow_html=True)
with col2:
    st.image("isg_avatar.png", width=150)


# API Key Kontrolü
api_key = None

try:
    # Tüm olası key varyasyonlarını dene
    possible_keys = ["GEMINI_API_KEY", "GOOGLE_API_KEY", "gemini_api_key", "google_api_key"]
    for k in possible_keys:
        if k in st.secrets:
            api_key = st.secrets[k]
            break
except Exception:
    pass

if not api_key:
    # Environment variable backup
    import os
    if os.getenv("GOOGLE_API_KEY"):
        api_key = os.getenv("GOOGLE_API_KEY")
    else:
        api_key = st.text_input("Google Gemini API Anahtarınızı Girin:", type="password")

if not api_key:
     st.warning("Devam etmek için Gemini API Key gereklidir.")
     st.stop()

# Otomatik Model Seçimi
try:
    genai.configure(api_key=api_key)
    # Mevcut modelleri listele
    available = list(genai.list_models())
    # generateContent destekleyen ilk modeli bul
    supported_models = [
        m.name for m in available 
        if 'generateContent' in m.supported_generation_methods
    ]
    
    # Öncelik sırası: Flash > Pro > Diğerleri
    selected_model = None
    if supported_models:
        # Önce flash var mı bak
        for m in supported_models:
            if "flash" in m.lower():
                selected_model = m
                break
        
        # Yoksa pro var mı bak
        if not selected_model:
            for m in supported_models:
                if "pro" in m.lower() and "vision" not in m.lower():
                    selected_model = m
                    break
        
        # Hiçbiri yoksa ilkini al
        if not selected_model:
            selected_model = supported_models[0]
            
    else:
        # Liste boş döndüyse fallback
        selected_model = "models/gemini-1.5-flash"

except Exception as e:
    # Listeleme hatası olursa (yetki vb.) fallback
    selected_model = "models/gemini-pro"
    # Debug için log (kullanıcıya gösterme)
    print(f"Model listeleme hatası: {e}")

with st.form("risk_form"):
    workplace = st.text_input("İşyeri / Sektör Tanımı:", placeholder="Örn: Mobilya Atölyesi, Demir Çelik Fabrikası, İnşaat Şantiyesi...")
    risk_count = st.slider("Oluşturulacak Risk Sayısı:", min_value=50, max_value=200, value=50, step=50)
    submitted = st.form_submit_button("Analizi Oluştur 🚀")


if submitted:
    if not api_key:
        st.error("Lütfen API Anahtarını kontrol edin.")
    elif not workplace:
        st.error("Lütfen bir işyeri tanımı girin.")
    else:
        # Random karikatür seçimi
        cartoons = [
            "isg_karikatur_1_1770212300830.png",
            "isg_karikatur_2_1770212343732.png",
            "isg_karikatur_3_1770212414232.png",
            "isg_karikatur_4_1770212438376.png",
            "isg_karikatur_5_1770212455882.png"
        ]
        selected_cartoon = random.choice(cartoons)
        
        # Karikatürü göster
        col_cartoon1, col_cartoon2, col_cartoon3 = st.columns([1, 2, 1])
        with col_cartoon2:
            st.image(selected_cartoon, use_container_width=True)
            st.markdown("<p style='text-align: center; color: #667eea; font-weight: 600; margin-top: 0.5rem;'>💡 İSG Bilgi: Güvenlik her zaman önceliğimizdir!</p>", unsafe_allow_html=True)
        
        st.markdown("---")
        
        status_text = st.empty()
        status_text.text("⏳ RİSK ANALİZİ OLUŞTURULUYOR...")
        progress_bar = st.progress(0)
        
        try:
            risks = fetch_risks_in_batches(api_key, selected_model, workplace, total_items=risk_count, batch_size=50, progress_bar=progress_bar, status_text=status_text)
            
            if risks:
                status_text.success(f"✅ {len(risks)} adet risk başarıyla analiz edildi!")
                
                # Excel Oluştur
                wb = create_excel(risks, workplace)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                # İndirme Butonu
                safe_name = "".join(c for c in workplace if c.isalnum() or c in (' ','-','_')).strip()
                st.download_button(
                    label="📥 Excel Dosyasını İndir",
                    data=output,
                    file_name=f"Risk_Analizi_{safe_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Önizleme (Opsiyonel)
                with st.expander("Sonuç Önizlemesi (İlk 5 Madde)"):
                    st.json(risks[:5])
                    
            else:
                st.error("Hiçbir risk verisi alınamadı. Lütfen tekrar deneyin.")
                
        except Exception as e:
            st.error(f"Beklenmeyen bir hata oluştu: {str(e)}")

# Footer (Sabit Alt Bilgi)
st.markdown("---")
st.markdown("""
    <div class='custom-footer'>
        <h3>⚡ İSG Risk Analiz Platformu</h3>
        <p>Bu Uygulama İş Güvenliği Uzmanı <strong>Fatih AKDENİZ</strong> tarafından geliştirilmiştir.</p>
        <p style='color: #718096; font-size: 0.85rem; margin-top: 1rem;'>📊 Fine-Kinney Metodolojisi ile Risk Değerlendirmesi</p>
    </div>
    """,
    unsafe_allow_html=True
)
