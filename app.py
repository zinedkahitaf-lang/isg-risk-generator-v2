from flask import Flask, render_template, request, send_file, jsonify
import openai
import json
import os
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)

# Fine Kinney Risk Seviyeleri ve Renkleri (Tosyalı Tablosu)
RISK_LEVELS = {
    "tolerans_gosterilemez": {"min": 400, "max": float('inf'), "color": "FFFF0000", "label": "Tolerans Gösterilemez Risk"},  # Parlak Kırmızı
    "esasli": {"min": 200, "max": 400, "color": "FF808080", "label": "Esaslı Risk"},  # Gri
    "onemli": {"min": 70, "max": 200, "color": "FF0070C0", "label": "Önemli Risk"},  # Mavi
    "olasi": {"min": 20, "max": 70, "color": "FFFFFF00", "label": "Olası Risk"},  # Sarı
    "onemsiz": {"min": 0, "max": 20, "color": "FF00B050", "label": "Önemsiz Risk"}  # Yeşil
}

def get_risk_level(score):
    """Fine Kinney risk seviyesi ve rengini döndür"""
    if score > 400:
        return RISK_LEVELS["tolerans_gosterilemez"]
    elif score > 200:
        return RISK_LEVELS["esasli"]
    elif score > 70:
        return RISK_LEVELS["onemli"]
    elif score > 20:
        return RISK_LEVELS["olasi"]
    else:
        return RISK_LEVELS["onemsiz"]

def fetch_risks_from_openai(api_key, workplace):
    """OpenAI API'den risk analizi al"""
    import httpx
    
    # httpx client ile OpenAI oluştur (proxy hatası çözümü)
    http_client = httpx.Client()
    client = openai.OpenAI(api_key=api_key, http_client=http_client)
    
    prompt = f"""
    Sen uzman bir İSG (İş Sağlığı ve Güvenliği) mühendisisin.
    Görev: '{workplace}' işyeri/sektörü için 50 adet detaylı risk değerlendirmesi yap.
    
    Fine Kinney Metodu değerleri:
    - Olasılık (O): 0.2, 0.5, 1, 3, 6, 10
    - Frekans (F): 0.5, 1, 2, 3, 6, 10
    - Şiddet (Ş): 1, 3, 7, 15, 40, 100
    
    Çıktı formatı: Sadece saf JSON array döndür. Markdown bloğu kullanma.
    Her obje şu anahtarları içermeli:
    - sira_no (1'den 50'ye kadar)
    - faaliyet_alani (Örn: Genel Yönetim, Üretim Alanı, Depo, vb.)
    - faaliyet_turu (Örn: Çalışma Ortamı, Makine Kullanımı, vb.)
    - tehlike_tanimi (Detaylı tehlike açıklaması)
    - risk_tanimi (Olası etki: yaralanma, ölüm, maddi hasar vb.)
    - olasilik (Fine Kinney değeri)
    - frekans (Fine Kinney değeri)
    - siddet (Fine Kinney değeri)
    - onlemler (DÖF - Düzeltici/Önleyici Faaliyetler, detaylı ve maddeler halinde)
    - sorumlu (Örn: İşveren & İSG Uzmanı, Şantiye Sorumlusu vb.)
    - sure (Aksiyon süresi: "Hemen", "1 Hafta", "1 Ay", "3 Ay" gibi göreceli süre - TARİH YAZMA!)
    - sonraki_olasilik (DÖF sonrası düşürülmüş değer)
    - sonraki_frekans (DÖF sonrası düşürülmüş değer)
    - sonraki_siddet (DÖF sonrası düşürülmüş değer)
    
    KRİTİK KURALLAR:
    1. DÖF sonrası Risk Skoru (O×F×Ş) KESİNLİKLE 70 veya altında olmalı!
    2. Yeşil (≤20): Önemsiz Risk
    3. Sarı (20-70): Olası Risk
    4. Profesyonel ve teknik bir dil kullan.
    5. "{workplace}" sektörüne/işyerine özel riskler üret.
    6. MUTLAKA 2-3 adet 400 üstü (Tolerans Gösterilemez) risk olmalı! Bu çok önemli.
    """

    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.7
    )
    
    content = response.choices[0].message.content.strip()
    
    # Markdown temizliği
    if content.startswith("```json"):
        content = content[7:]
    if content.startswith("```"):
        content = content[3:]
    if content.endswith("```"):
        content = content[:-3]
    
    return json.loads(content)

def create_excel(risk_data, workplace):
    """Excel dosyası oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Değerlendirme"
    
    # Başlıklar
    headers = [
        "Sıra No", "Faaliyet Alanı", "Faaliyet Türü", 
        "Tehlike Tanımı", "Risk Tanımı (Olası Etki)",
        "O", "F", "Ş", "R", "Riskin Tanımı",
        "Planlanan Faaliyetler / DÖF", "Sorumlu", "Süre",
        "Sonraki O", "Sonraki F", "Sonraki Ş", "Sonraki R", "Sonraki Riskin Tanımı"
    ]
    
    ws.append(headers)
    
    # Stiller
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Başlık stilleri
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Veri ekle
    for item in risk_data:
        o = float(item.get('olasilik', 1))
        f = float(item.get('frekans', 1))
        s = float(item.get('siddet', 1))
        current_score = o * f * s
        current_level = get_risk_level(current_score)
        
        so = float(item.get('sonraki_olasilik', 0.2))
        sf = float(item.get('sonraki_frekans', 1))
        ss = float(item.get('sonraki_siddet', 1))
        next_score = so * sf * ss
        next_level = get_risk_level(next_score)
        
        # Önlemleri string'e çevir (AI bazen liste döndürüyor)
        onlemler = item.get('onlemler', '')
        if isinstance(onlemler, list):
            onlemler = '\n'.join([f"• {o}" for o in onlemler])
        
        # Diğer alanları da kontrol et
        tehlike = item.get('tehlike_tanimi', '')
        if isinstance(tehlike, list):
            tehlike = ' '.join(tehlike)
        
        risk = item.get('risk_tanimi', '')
        if isinstance(risk, list):
            risk = ' '.join(risk)
        
        row = [
            item.get('sira_no'),
            item.get('faaliyet_alani'),
            item.get('faaliyet_turu'),
            tehlike,
            risk,
            o, f, s, current_score, current_level["label"],
            onlemler,
            item.get('sorumlu'),
            item.get('sure'),
            so, sf, ss, next_score, next_level["label"]
        ]
        ws.append(row)
    
    # Satır stilleri ve renklendirme
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Mevcut Risk Skoru renklendirme (9. sütun - R)
        score_cell = ws.cell(row=row_idx, column=9)
        try:
            val = float(score_cell.value)
            level = get_risk_level(val)
            score_cell.fill = PatternFill(start_color=level["color"], end_color=level["color"], fill_type="solid")
            if level["color"] in ["FF0070C0", "FF808080", "FFFF0000"]:  # Mavi, Gri, Kırmızı
                score_cell.font = Font(color="FFFFFF", bold=True)
        except:
            pass
        
        # Sonraki Risk Skoru renklendirme (17. sütun - Sonraki R)
        next_score_cell = ws.cell(row=row_idx, column=17)
        try:
            val = float(next_score_cell.value)
            level = get_risk_level(val)
            next_score_cell.fill = PatternFill(start_color=level["color"], end_color=level["color"], fill_type="solid")
        except:
            pass
    
    # Sütun genişlikleri
    column_widths = {
        'A': 8, 'B': 18, 'C': 18, 'D': 35, 'E': 30,
        'F': 6, 'G': 6, 'H': 6, 'I': 8, 'J': 22,
        'K': 50, 'L': 25, 'M': 12,
        'N': 8, 'O': 8, 'P': 8, 'Q': 10, 'R': 22
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Satır yüksekliği
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 60
    
    return wb

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate():
    try:
        data = request.json
        workplace = data.get('workplace')
        
        # API key'i environment variable'dan al (güvenlik için)
        api_key = os.environ.get('OPENAI_API_KEY')
        
        if not api_key:
            return jsonify({"error": "Sunucu yapılandırma hatası: API Key bulunamadı"}), 500
        
        if not workplace:
            return jsonify({"error": "İşyeri bilgisi gerekli"}), 400
        
        # Risk analizi al
        risks = fetch_risks_from_openai(api_key, workplace)
        
        # Excel oluştur
        wb = create_excel(risks, workplace)
        
        # Belleğe kaydet
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Dosya adı
        safe_name = "".join(c for c in workplace if c.isalnum() or c in (' ', '-', '_')).strip()
        filename = f"ISG_Risk_Analizi_{safe_name}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except json.JSONDecodeError as e:
        return jsonify({"error": f"AI yanıtı işlenemedi: {str(e)}"}), 500
    except openai.APIError as e:
        return jsonify({"error": f"OpenAI Hatası: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": f"Beklenmeyen hata: {str(e)}"}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
